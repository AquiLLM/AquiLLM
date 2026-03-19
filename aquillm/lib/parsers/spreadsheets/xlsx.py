"""
Excel XLSX spreadsheet parsing.
"""

import io


def extract_xlsx_text(data: bytes) -> str:
    """Extract text content from XLSX bytes."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise ValueError("openpyxl is required for .xlsx extraction.") from exc
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                lines.append(", ".join(values))
    return "\n".join(lines)


__all__ = ['extract_xlsx_text']
